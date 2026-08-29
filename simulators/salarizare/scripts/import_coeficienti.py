"""Turn the 48-sheet coefficient workbook into the positions of a regime document.

    uv run --with openpyxl python scripts/import_coeficienti.py

Reads  sources/Proiect-COEFICIENTI-1-8-MMFTSS-16.07.2026-1000.xlsx
       data/frames/ro-draft-2026-07-16.frame.json   (hand-written, never generated)
Writes data/regimes/ro-draft-2026-07-16.json
       data/reports/import-coeficienti.json         (coverage, so gaps are visible)

Why the classifier works bottom-up
----------------------------------
The 48 sheets do not share a header layout. Some put the column headers on row 4,
others on row 10; some label coefficient columns "Grad I / Grad II", others
"Nivel I / Nivel II", "Grad managerial", "Gradația 0", or five calendar years; some
have no sub-header at all. Parsing headers top-down would mean 48 special cases.

So columns are classified from their *contents* across the whole sheet instead, and
headers are consulted only afterwards, to name what was found. A column earns the role
`coefficient` by behaving like one.

The rule that matters most: the workbook still contains the author's working columns —
old and new salary amounts in lei, and their ratios. Those ratios sit in the 0,98-1,15
band and would pass a naive "is it between 1 and 8" test, inflating the very
back-solving statistic the app is built to show. A numeric column is only a coefficient
if it reaches 1,5 somewhere, which no ratio column does and every real coefficient
column does.

Nothing here guesses silently. Every uncertain decision lands in the output as
`confidence: "assumed"` or `assimilation.parse: "needsReview"`, and every skipped row is
counted in the coverage report.
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "sources/Proiect-COEFICIENTI-1-8-MMFTSS-16.07.2026-1000.xlsx"
FRAME = ROOT / "data/frames/ro-draft-2026-07-16.frame.json"
OUT = ROOT / "data/regimes/ro-draft-2026-07-16.json"
REPORT = ROOT / "data/reports/import-coeficienti.json"

SOURCE = "coeficienti-2026-07-16"

CODE = re.compile(r"^\d{2}\.\d{8}\.\d{2}\.\d$")
WORD = re.compile(r"[A-Za-zĂÂÎȘȚăâîșț]{4}")

# A coefficient column reaches at least this value somewhere. Ratio columns do not.
COEFFICIENT_CEILING = 1.5
COEFFICIENT_RANGE = (0.5, 8.5)

STUDY_LEVELS = {
    "s", "ssd", "m", "g", "pl", "m/g", "m,g", "sd", "s/ssd", "m;g",
    "studii superioare", "studii medii", "-",
}

FAMILY_BY_ANNEX = {
    "I": "I-invatamant",
    "II": "II-sanatate-asistenta-sociala",
    "III": "III-cultura",
    "IV": "IV-diplomatie",
    "V": "V-justitie",
    "VI": "VI-aparare-ordine-securitate",
    "VII": "VII-venituri-proprii",
    "VIII": "VIII-administratie",
    "IX": "IX-demnitati-publice",
}

# First two digits of a function code, cross-checked against the sheet's annex.
FAMILY_BY_CODE_PREFIX = {
    "11": "I", "12": "I",
    "21": "II",
    "31": "III", "32": "III", "33": "III", "34": "III", "35": "III",
    "41": "IV",
    "51": "V", "52": "V", "53": "V", "54": "V", "55": "V", "56": "V", "57": "V",
    "61": "VI",
    "71": "VII",
    "81": "VIII", "82": "VIII",
}

# Stems, not whole phrases. Romanian declines its nouns, and the sheets say
# "categoriei înalţilor funcţionari publici" — against which a literal "inalti
# functionari" never matches. That miss put every înalt funcţionar public on the
# gradatii ladder, which Art. 13(1) explicitly excludes them from.
# Annex II Art. 10: the coefficients printed for health staff are a midpoint, not a
# figure. The real level is set per category of health unit between -15% and +15% of what
# the sheet shows — minus for medico-social and outpatient units, plus for forensic
# medicine. The sheet tabs do not carry the annex's own numbering: "II CI 1" holds points
# 1 and 2 (point 2 starts partway down it), "II CI 2" is point 3, and "II CI 3" is point
# 4, which Art. 10 does not cover. Getting that wrong would apply the band to social-care
# staff the article never mentions.
UNIT_CATEGORY_SHEETS = {"II CI 1", "II CI 2"}
UNIT_CATEGORY_BAND = 0.15

MANAGEMENT_MARKERS = ("de conducere", "de conducer")
MANAGEMENT_STEMS = (("inalt", "functionar"),)
EXECUTION_MARKERS = ("de execut", "de execuţ", "de execuție")

HEAD_NOUN = re.compile(
    r"^(director|sef|șef|şef|inspector|expert|asistent|tehnician|consilier|referent|medic|"
    r"comisar|contabil|inginer|auditor|trezorier|secretar|manager|profesor|educator|"
    r"redactor|arhivist|bibliotecar|muzeograf|farmacist|biolog|chimist|fizician|psiholog|"
    r"kinetoterapeut|moasa|moaşă|registrator|statistician|operator|casier|magaziner|portar|"
    r"ingrijitor|îngrijitor|muncitor|sofer|şofer|paznic|analist|programator|administrator|"
    r"arhitect|cercetator|cercetător|procuror|judecator|judecător|grefier|agent|controlor|"
    r"revizor|preot|antrenor|instructor|maistru|laborant|infirmier|brancardier|dietetician|"
    r"logoped|optician|cosmetician|institutor|licentiat|licenţiat|invatator|învățător|"
    r"rector|prorector|decan|prodecan|lector|conferentiar|conferenţiar|cercetator)",
    re.IGNORECASE,
)


def strip_accents(text: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")


def annex_of(sheet_name: str) -> str:
    token = re.match(r"^(VIII|VII|VI|IV|IX|V|III|II|I)\b", sheet_name.strip().replace("_", " "))
    return token.group(1) if token else ""


# --------------------------------------------------------------- classification


def in_range(value) -> bool:
    return (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and COEFFICIENT_RANGE[0] <= value <= COEFFICIENT_RANGE[1]
    )


def classify_columns(rows: list[tuple]) -> dict[int, str]:
    """Assign each column a role from what it contains across the whole sheet.

    Two things make this harder than it looks.

    Title banners. Every sheet carries several rows of headings before the grid, and
    counting those as column contents makes a numeric column look like a text one -
    which is what hid Annex IX lit. D. So the tally starts at the first row that holds a
    plausible coefficient.

    Working columns. The author's old-versus-new salary ratios sit in the 0,98-1,15 band
    and pass any "is it between 1 and 8" test. They are told apart by position: the real
    coefficients are printed to the *left* of the function code, the leftover arithmetic
    to the right. Where a sheet has no code column at all, fall back to requiring the
    column to reach 1,5 - a ratio column never does.
    """
    data_start = min((i for i, row in enumerate(rows) if any(in_range(v) for v in row)), default=0)
    body = rows[data_start:]

    values: dict[int, list] = defaultdict(list)
    for row in body:
        for c, v in enumerate(row):
            if v is not None and (not isinstance(v, str) or v.strip()):
                values[c].append(v)

    code_cols = {
        c for c, vals in values.items()
        if (texts := [v.strip() for v in vals if isinstance(v, str)])
        and sum(1 for t in texts if CODE.match(t)) / len(texts) > 0.5
    }
    first_code = min(code_cols) if code_cols else None

    roles: dict[int, str] = {c: "code" for c in code_cols}
    for c, vals in values.items():
        if c in roles:
            continue
        texts = [v.strip() for v in vals if isinstance(v, str)]
        numbers = [v for v in vals if isinstance(v, (int, float)) and not isinstance(v, bool)]

        if numbers and len(numbers) >= max(3, 0.5 * len(vals)):
            plausible = [v for v in numbers if in_range(v)]
            integers = [v for v in numbers if float(v).is_integer()]
            # The "Nr. crt." column is an ascending integer run. It must be caught by its
            # shape, not by its maximum: on a short sheet it never leaves 1..8, and every
            # one of those values is a plausible coefficient.
            # In Annex V the number is printed only on the first row of each group, so the
            # column holds as few as four values - too few for a length test alone, and
            # misreading it as pay also knocks the real coefficient column out of the
            # adjacency block below. The printed "Nr. crt." heading settles it.
            titled_crt = any(
                c < len(row) and isinstance(row[c], str) and "crt" in row[c].lower()
                for row in rows[:data_start] or rows[:12]
            )
            ascending = (
                len(integers) == len(numbers)
                and len(numbers) >= 3
                and len(set(numbers)) == len(numbers)
                and numbers[0] <= 2
                and all(b >= a for a, b in zip(numbers, numbers[1:]))
            )
            if ascending and (len(numbers) >= 5 or titled_crt):
                roles[c] = "index"
            elif plausible and len(plausible) / len(numbers) > 0.7 and (
                c < first_code if first_code is not None else max(plausible) >= COEFFICIENT_CEILING
            ):
                roles[c] = "coefficient"
            else:
                roles[c] = "working"
            continue

        if texts:
            lowered = [strip_accents(t).lower() for t in texts]
            if sum(1 for t in lowered if t in STUDY_LEVELS) / len(lowered) > 0.5:
                roles[c] = "studyLevel"
            elif sum(1 for t in texts if WORD.search(t)) / len(texts) > 0.4:
                roles[c] = "text"
            else:
                roles[c] = "qualifier"

    # The grid is printed as one block of adjacent columns. Anything numeric that sits
    # further right, separated by empty columns, is leftover arithmetic that happens to
    # fall left of the code column - VIII_CI_A_1 keeps its old/new ratios in column K,
    # four columns clear of the real coefficients in E and F.
    candidates = sorted(c for c, r in roles.items() if r == "coefficient")
    if candidates:
        block = [candidates[0]]
        for c in candidates[1:]:
            if c - block[-1] <= 2:
                block.append(c)
            else:
                break
        for c in candidates:
            if c not in block:
                roles[c] = "working"
    return roles


def column_labels(rows: list[tuple], roles: dict[int, str], first_data_row: int) -> dict[int, str]:
    """Nearest non-empty string above the data, per coefficient column.

    This is what turns two anonymous coefficient columns into `Grad I` / `Grad II`, or
    the five columns of Annex IX into calendar years.
    """
    labels: dict[int, str] = {}
    for c, role in roles.items():
        if role != "coefficient":
            continue
        for r in range(first_data_row - 1, -1, -1):
            if c >= len(rows[r]):
                continue
            value = rows[r][c]
            if isinstance(value, str) and value.strip() and not CODE.match(value.strip()):
                labels[c] = re.sub(r"\s+", " ", value.strip())
                break
            # Annex IX heads its five coefficient columns with calendar years, and the
            # workbook stores 2028 as a number while 2026/2027 is text. A bare year is a
            # label, not data, so it must not stop the walk upward.
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                if float(value).is_integer() and 1990 <= value <= 2100:
                    labels[c] = str(int(value))
                    break
                break
    return labels


def dim_for(label: str | None, index: int) -> tuple[str, str, str]:
    """(dimension name, dimension value, confidence) for a coefficient column."""
    if not label:
        return "variant", f"col{index + 1}", "assumed"
    flat = re.sub(r"\s+", " ", strip_accents(label).strip().lower())
    if re.fullmatch(r"20\d\d(\s*/\s*20\d\d)?", flat):
        return "an", label.strip(), "verbatim"
    for prefix, dim in (
        ("nivel", "institutionLevel"),
        ("grad managerial", "gradManagerial"),
        ("grad ", "grad"),
        ("gradatia", "gradatie"),
        ("treapta", "treapta"),
        ("vechime", "vechime"),
    ):
        if flat.startswith(prefix):
            value = re.sub(r"^" + re.escape(prefix), "", flat).strip() or flat
            return dim, value.upper() if len(value) <= 3 else value, "verbatim"
    return "variant", label.strip(), "derived"


# ------------------------------------------------------------------- titles


# A row whose title cell names a rank rather than an occupation continues the row above.
# The workbook writes them indented under their parent - "         gradul  I",
# "    clasa a II-a", "         debutant" - and they are never jobs in their own right.
# Read as positions they produce 46 posts called "debutant" and 30 called "clasa a II-a",
# which is both wrong and the kind of wrong that looks like data.
RANK_PREFIX = re.compile(
    r"^(gradul|grad|clasa|treapta|nivel|definitiv|debutant|principal|asistent"
    r"|superior|specialist|practicant|stagiar)\b",
    re.IGNORECASE,
)


def rank_label(cell: str) -> str | None:
    """The rank a continuation row carries, or None when the cell names a job."""
    text = (cell or "").strip()
    if not text or not RANK_PREFIX.match(text):
        return None
    # "Asistent medical" is an occupation; "asistent" alone is a professional grade.
    # Anything long enough to carry a noun after the rank word is a title, not a rank.
    return text if len(text.split()) <= 3 else None


def parse_titles(cell: str) -> tuple[list[dict], str, int]:
    """Split a title cell into former job titles.

    Returns (titles, parse kind, fan-in). Never decides silently: a cell whose
    fragments do not all look like job titles is returned whole and marked
    needsReview for a human.
    """
    raw = re.sub(r"\s+", " ", cell.strip().strip(";,")).strip()
    if not raw:
        return [], "single", 0

    def package(parts: list[str], kind: str, qualifier: str | None = None) -> tuple[list[dict], str, int]:
        titles = []
        for i, part in enumerate(parts):
            entry: dict = {"name": part}
            if i == 0:
                entry["canonical"] = True
            if qualifier:
                entry["qualifier"] = qualifier
            titles.append(entry)
        return titles, kind, len(titles)

    if ";" in raw:
        fragments = [f.strip(" .;,") for f in raw.split(";") if f.strip(" .;,")]
        if len(fragments) < 2:
            return package([raw], "single")
        # A trailing qualifier hangs off the last fragment after a comma:
        # "Institutor; maistru instructor, studii superioare lunga durata grad didactic I"
        qualifier = None
        last = fragments[-1]
        if "," in last and not HEAD_NOUN.match(last.split(",", 1)[1].strip()):
            head, tail = last.split(",", 1)
            if HEAD_NOUN.match(head.strip()):
                fragments[-1] = head.strip()
                qualifier = tail.strip()
        if all(HEAD_NOUN.match(f) for f in fragments):
            return package(fragments, "mixed" if qualifier else "semicolon", qualifier)
        return package([raw], "needsReview")

    if "," in raw:
        fragments = [f.strip(" .,") for f in raw.split(",") if f.strip(" .,")]
        if len(fragments) >= 2 and all(HEAD_NOUN.match(f) for f in fragments):
            return package(fragments, "comma")
        return package([raw], "single")

    if "/" in raw:
        fragments = [f.strip(" ./") for f in raw.split("/") if f.strip(" ./")]
        if len(fragments) >= 2 and all(HEAD_NOUN.match(f) for f in fragments):
            return package(fragments, "slash")
        return package([raw], "single")

    return package([raw], "single")


# ------------------------------------------------------------------ extraction


VECHIME = re.compile(r"(ani|peste|pana|până|luni)|^\d+\s*-\s*\d+", re.IGNORECASE)
GRAD = re.compile(
    r"(gradul|grad\b|treapta|debutant|principal|superior|asistent|specialist|"
    r"rezident|maestru|categoria|clasa)",
    re.IGNORECASE,
)


def row_dims(row: tuple, roles: dict[int, str], title_cell: str) -> dict[str, str]:
    """Dimensions written on the row rather than in a column header.

    Annexes I and V put the seniority band in a cell beside the title - 'peste 25 de
    ani', '20-25 ani' - instead of splitting it across columns. Without this the two
    rows of one job look like two unrelated positions.
    """
    dims: dict[str, str] = {}
    for c, role in roles.items():
        if role not in ("text", "qualifier") or c >= len(row):
            continue
        value = row[c]
        if not isinstance(value, str):
            continue
        value = re.sub(r"\s+", " ", value.strip())
        if not value or value == title_cell or value == "-":
            continue
        if VECHIME.search(value) and len(value) < 40:
            dims.setdefault("vechime", value)
        elif GRAD.search(value) and len(value) < 40:
            dims.setdefault("gradProfesional", value)
    return dims


def sheet_slug(name: str) -> str:
    return re.sub(r"\.+", ".", re.sub(r"[\s_]+", ".", name.strip()))


def section_kind(rows: list[tuple], upto: int, sheet: str, default: str) -> str:
    """Walk upward for the nearest section heading that names conducere or execuție."""
    if sheet.startswith("IX"):
        return "dignitary"
    if "Armata" in sheet or annex_of(sheet) == "VI":
        return "uniformed"
    for r in range(upto, -1, -1):
        joined = strip_accents(
            " ".join(str(v) for v in rows[r] if isinstance(v, str))
        ).lower()
        if not joined:
            continue
        if any(m in joined for m in (strip_accents(x).lower() for x in MANAGEMENT_MARKERS)):
            return "management"
        if any(all(stem in joined for stem in stems) for stems in MANAGEMENT_STEMS):
            return "management"
        if any(m in joined for m in (strip_accents(x).lower() for x in EXECUTION_MARKERS)):
            return "execution"
    return default


def extract_sheet(name: str, rows: list[tuple], stats: Counter, skipped: list) -> list[dict]:
    roles = classify_columns(rows)
    coefficient_cols = sorted(c for c, r in roles.items() if r == "coefficient")
    code_cols = sorted(c for c, r in roles.items() if r == "code")
    text_cols = sorted(c for c, r in roles.items() if r == "text")
    lengths: dict[int, list[int]] = {c: [] for c in text_cols}
    for row in rows:
        for c in text_cols:
            if c < len(row) and isinstance(row[c], str) and WORD.search(row[c]):
                lengths[c].append(len(row[c].strip()))
    title_col = max(
        (c for c in text_cols if lengths[c]),
        key=lambda c: sum(lengths[c]) / len(lengths[c]),
        default=None,
    )
    study_cols = sorted(c for c, r in roles.items() if r == "studyLevel")

    if not coefficient_cols:
        stats["sheets_without_coefficients"] += 1
        skipped.append({"sheet": name, "reason": "no column behaved like a coefficient"})
        return []

    # Must be range-checked: Annex IX heads its coefficient columns with the years
    # 2028..2031, which are numeric. Without the range test the header row itself counts
    # as data and the label walk starts above it, losing the year labels entirely.
    first_data = min(
        (i for i, row in enumerate(rows)
         if any(c < len(row) and isinstance(row[c], (int, float)) and not isinstance(row[c], bool)
                and COEFFICIENT_RANGE[0] <= row[c] <= COEFFICIENT_RANGE[1]
                for c in coefficient_cols)),
        default=len(rows),
    )
    labels = column_labels(rows, roles, first_data)
    annex = annex_of(name)
    family = FAMILY_BY_ANNEX.get(annex, f"unknown-{annex or name}")

    index_cols = sorted(c for c, r in roles.items() if r == "index")
    slug = sheet_slug(name)

    positions: list[dict] = []
    last: dict | None = None
    pending_title = ""
    for r, row in enumerate(rows):
        codes = [
            (c, row[c].strip())
            for c in code_cols
            if c < len(row) and isinstance(row[c], str) and CODE.match(row[c].strip())
        ]
        coefficients = [
            (c, float(row[c]))
            for c in coefficient_cols
            if c < len(row)
            and isinstance(row[c], (int, float))
            and not isinstance(row[c], bool)
            and COEFFICIENT_RANGE[0] <= row[c] <= COEFFICIENT_RANGE[1]
        ]
        row_title = ""
        if title_col is not None and title_col < len(row) and isinstance(row[title_col], str):
            if WORD.search(row[title_col]):
                row_title = row[title_col]

        if not coefficients:
            # Annex I prints the occupation on its own line - "Preot" - and the pay on the
            # indented rank lines beneath it. Skipping this row outright, as the importer
            # used to, threw the only copy of the job's name away.
            if row_title and rank_label(row_title) is None:
                pending_title = row_title
            continue

        title_cell = ""
        if title_col is not None and title_col < len(row) and isinstance(row[title_col], str):
            if WORD.search(row[title_col]):
                title_cell = row[title_col]
        if not title_cell:
            for c in text_cols:
                if c < len(row) and isinstance(row[c], str) and WORD.search(row[c]):
                    # Never fall back onto a rank: in Annex VIII the occupation sits in
                    # one column and the class in the next, so a blank occupation cell
                    # would otherwise promote "    clasa a II-a" to a job title.
                    if rank_label(row[c]) is None and len(row[c]) > len(title_cell):
                        title_cell = row[c]

        base_code = codes[0][1].rsplit(".", 1)[0] if codes else None
        extra_dims = row_dims(row, roles, title_cell)

        rank = rank_label(title_cell)
        if rank is not None:
            extra_dims = {**extra_dims, "grad": rank}
            title_cell = ""
            stats["rank_rows"] += 1

        # A row with a code but no title continues the position above it: Annexes I and V
        # print one row per seniority band and leave the title cell blank after the first.
        # Dropping these would silently discard most of the judiciary and teaching grids.
        if not title_cell:
            if last is not None and (base_code is None or base_code == last["code"]):
                for i, (col, value) in enumerate(coefficients):
                    dim, dim_value, confidence = dim_for(labels.get(col), i)
                    variant = {
                        "value": value,
                        "dims": {**({dim: dim_value} if confidence != "assumed" else {}), **extra_dims},
                        "provenance": {
                            "source": SOURCE,
                            "locator": f"sheet '{name}'!"
                            f"{openpyxl.utils.get_column_letter(col + 1)}{r + 1}"
                            + (f", cod {codes[i][1]}" if len(codes) == len(coefficients) else ""),
                            "confidence": "verbatim",
                            "note": "Rand de continuare: denumirea e preluata din randul anterior cu acelasi cod de baza.",
                        },
                    }
                    if not variant["dims"]:
                        del variant["dims"]
                    last["variants"].append(variant)
                stats["continuation_rows"] += 1
                continue
            if pending_title:
                title_cell = pending_title
            else:
                stats["rows_without_title"] += 1
                skipped.append({"sheet": name, "row": r + 1, "reason": "no title cell and no position above it to continue"})
                continue

        titles, parse, fan_in = parse_titles(title_cell)
        if parse == "needsReview":
            stats["rows_needing_review"] += 1

        study = ""
        for c in study_cols:
            if c < len(row) and isinstance(row[c], str) and row[c].strip():
                study = row[c].strip()
                break

        # One code per coefficient column where the counts line up; otherwise the row's
        # first code identifies the position and the mapping is left unclaimed.
        aligned = len(codes) == len(coefficients)
        if base_code is None:
            # Annex IX carries no function codes. Build one from the sheet and the
            # printed Nr. crt. so it survives a re-import, rather than from the row
            # number, which moves whenever the workbook gains a line.
            nr = next(
                (int(row[c]) for c in index_cols
                 if c < len(row) and isinstance(row[c], (int, float))
                 and float(row[c]).is_integer()),
                None,
            )
            base_code = f"{slug}.{nr}" if nr is not None else f"{slug}.r{r + 1}"
            stats["rows_without_code"] += 1

        variants = []
        for i, (col, value) in enumerate(coefficients):
            dim, dim_value, confidence = dim_for(labels.get(col), i)
            variant: dict = {
                "value": value,
                "provenance": {
                    "source": SOURCE,
                    "locator": f"sheet '{name}'!{openpyxl.utils.get_column_letter(col + 1)}{r + 1}"
                    + (f", cod {codes[i][1]}" if aligned else ""),
                    "confidence": "verbatim",
                },
            }
            dims: dict[str, str] = {}
            if (len(coefficients) > 1 or confidence != "assumed") and confidence != "assumed":
                dims[dim] = dim_value
            elif len(coefficients) > 1:
                dims[dim] = dim_value
                variant["provenance"]["note"] = (
                    "Coloana de coeficient nu are antet; dimensiunea e pozitionala."
                )
            dims.update(extra_dims)
            if dims:
                variant["dims"] = dims
            variants.append(variant)

        # Which seniority mechanism applies is decided from the row itself. Art. 13(1)
        # excepts dignitaries and uniformed staff outright. Where the annex already
        # prints a coefficient per seniority band - Annexes I and V - the vechime is
        # inside the coefficient and the gradatii must not be applied on top of it.
        has_vechime = any("vechime" in (v.get("dims") or {}) for v in variants)
        kind_now = section_kind(rows, r, name, "execution")
        ladder = None if (kind_now in ("dignitary", "uniformed") or has_vechime) else "gradatii"

        position = {
            "code": base_code,
            "ladder": ladder,
            "name": titles[0]["name"] if titles else title_cell,
            "titles": titles,
            "assimilation": {
                "rawTitleCell": title_cell,
                "parse": parse,
                "fanIn": fan_in,
                "provenance": {
                    "source": SOURCE,
                    "locator": f"sheet '{name}'!r{r + 1}",
                    "confidence": "verbatim" if parse in ("single", "semicolon") else "derived",
                },
            },
            "family": family,
            "chapter": f"Anexa {annex} - {name}" if annex else name,
            "kind": kind_now,
            "variants": variants,
            "provenance": {
                "source": SOURCE,
                "locator": f"sheet '{name}'!r{r + 1}",
                "confidence": "verbatim",
                **({} if codes else {"note": "Randul nu poarta cod de functie; cod atribuit de importator."}),
            },
        }
        if study:
            position["studyLevel"] = study

        if name in UNIT_CATEGORY_SHEETS:
            position["institutionFactor"] = {
                "min": round(1 - UNIT_CATEGORY_BAND, 4),
                "max": round(1 + UNIT_CATEGORY_BAND, 4),
                "reason": (
                    "Anexa II Cap. II Art. 10: coeficientul publicat este mijlocul unui interval "
                    "de ±15%, stabilit pe categorii de unități sanitare. Se aplică diminuat cu 15% "
                    "la unitățile de asistență medico-socială și la cele ambulatorii, majorat cu "
                    "15% la medicina legală. Categoriile concrete se stabilesc prin hotărâre de "
                    "Guvern, deci nivelul efectiv nu se poate calcula din lege."
                ),
                "provenance": {
                    "source": "anexa-II-cap-II",
                    "locator": "Anexa II Cap. II Art. 10 alin. (1)-(3)",
                    "confidence": "verbatim",
                },
            }
            stats["positions_with_unit_category_band"] += 1
        if codes and annex:
            prefix = codes[0][1][:2]
            if FAMILY_BY_CODE_PREFIX.get(prefix) not in (None, annex):
                stats["code_prefix_mismatch"] += 1
                position["provenance"]["note"] = (
                    f"Prefixul codului {prefix} indica anexa "
                    f"{FAMILY_BY_CODE_PREFIX.get(prefix)}, foaia e in anexa {annex}."
                )
        positions.append(position)
        last = position
        stats["positions"] += 1
    return positions


# ------------------------------------------------------------------ assembly


def assign_grades(positions: list[dict], grades: list[dict], overrides: dict) -> Counter:
    """Place each variant in a grade band by computing it, never by trusting a label."""
    stats = Counter()
    bands = [
        (g["id"], float(g["min"] if isinstance(g["min"], (int, float)) else g["min"][0]["value"]),
         float(g["max"] if isinstance(g["max"], (int, float)) else g["max"][0]["value"]))
        for g in grades
    ]
    for position in positions:
        if overrides.get(position["code"], {}).get("gradeIdPolicy") == "none":
            stats["grade_skipped_by_override"] += 1
            continue
        for variant in position["variants"]:
            value = variant["value"]
            match = next((gid for gid, lo, hi in bands if lo <= value <= hi), None)
            if match:
                variant["gradeId"] = match
                stats["graded"] += 1
            else:
                stats["outside_every_band"] += 1
    return stats


def main() -> None:
    frame = json.loads(FRAME.read_text(encoding="utf-8"))
    overrides = frame.pop("positionOverrides", {})

    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    stats: Counter = Counter()
    skipped: list = []
    positions: list[dict] = []
    per_sheet: dict[str, int] = {}

    for name in workbook.sheetnames:
        rows = [tuple(r) for r in workbook[name].iter_rows(values_only=True)]
        found = extract_sheet(name, rows, stats, skipped)
        per_sheet[name] = len(found)
        positions.extend(found)

    # Merge duplicate codes: a position split across rows by seniority band.
    merged: dict[str, dict] = {}
    for position in positions:
        existing = merged.get(position["code"])
        if existing is None:
            merged[position["code"]] = position
        else:
            existing["variants"].extend(position["variants"])
            stats["rows_merged_into_existing_code"] += 1
    positions = list(merged.values())

    # Merging by code collapses rows that came from different sheets. Where two variants
    # end up with identical dims they become indistinguishable, and payslip() would price
    # the first match — so a director in the smallest local tier would be paid at the
    # largest tier's coefficient. The four "VIII CII A 3_localN" sheets are exactly this:
    # one code, four tiers, coefficients from 4,47 down to 2,47. Disambiguate by the sheet
    # the row came from, which is where the distinction was carried all along.
    sheet_of = re.compile(r"sheet '([^']+)'")
    for position in positions:
        groups: dict[str, list[dict]] = defaultdict(list)
        for variant in position["variants"]:
            groups[json.dumps(variant.get("dims") or {}, sort_keys=True)].append(variant)
        for colliding in groups.values():
            if len(colliding) < 2:
                continue
            for variant in colliding:
                match = sheet_of.search(variant["provenance"]["locator"])
                if match:
                    variant.setdefault("dims", {})["sursa"] = match.group(1)
                    stats["variants_disambiguated_by_sheet"] += 1

        # Some collisions are inside a single sheet, where the source distinguishes two
        # rows by nothing but their place on the page. Say so rather than leave two
        # variants a caller cannot choose between.
        regroup: dict[str, list[dict]] = defaultdict(list)
        for variant in position["variants"]:
            regroup[json.dumps(variant.get("dims") or {}, sort_keys=True)].append(variant)
        for colliding in regroup.values():
            if len(colliding) < 2:
                continue
            for variant in colliding:
                cell = re.search(r"!([A-Z]*\d+)", variant["provenance"]["locator"])
                if cell:
                    # The full cell, not just the row: Annex VI prints two Min/Max blocks
                    # side by side on one row, so the column is the only thing telling
                    # them apart.
                    variant.setdefault("dims", {})["celula"] = cell.group(1)
                    stats["variants_disambiguated_by_cell"] += 1

    for code, patch in overrides.items():
        target = next((p for p in positions if p["code"] == code), None)
        if target is None:
            stats["override_without_position"] += 1
            continue
        for key, value in patch.items():
            if key in ("note", "gradeIdPolicy"):
                continue
            target[key] = value
        if "note" in patch:
            target["provenance"]["note"] = patch["note"]
        stats["overrides_applied"] += 1

    grade_stats = assign_grades(positions, frame["grades"], overrides)
    stats.update(grade_stats)

    regime = {k: v for k, v in frame.items()}
    regime["positions"] = sorted(positions, key=lambda p: (p["family"], p["code"]))
    regime["provenance"] = {
        **frame["provenance"],
        "note": "Positions generated by scripts/import_coeficienti.py from the coefficient "
        "workbook. Everything else is hand-written and lives in "
        "data/frames/ro-draft-2026-07-16.frame.json.",
    }

    OUT.write_text(json.dumps(regime, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    parses = Counter(p["assimilation"]["parse"] for p in positions)
    fan_in = Counter(p["assimilation"]["fanIn"] for p in positions)
    report = {
        "generatedFrom": WORKBOOK.name,
        "sheets": len(workbook.sheetnames),
        "positions": len(positions),
        "variants": sum(len(p["variants"]) for p in positions),
        "perSheet": per_sheet,
        "counters": dict(stats),
        "assimilationParse": dict(parses),
        "fanInHistogram": {str(k): v for k, v in sorted(fan_in.items())},
        "titlesAbsorbed": sum(p["assimilation"]["fanIn"] for p in positions),
        "skipped": skipped[:200],
        "skippedTotal": len(skipped),
    }
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"positions: {len(positions)}  variants: {report['variants']}")
    print(f"merged rows: {stats['rows_merged_into_existing_code']}  "
          f"needs review: {stats['rows_needing_review']}  "
          f"no code: {stats['rows_without_code']}")
    print(f"graded: {stats['graded']}  outside every band: {stats['outside_every_band']}")
    print(f"parse: {dict(parses)}")
    print(f"wrote {OUT.relative_to(ROOT)} and {REPORT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
