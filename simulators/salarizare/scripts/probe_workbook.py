"""Read-only probes over the coefficient workbook.

This is not the importer. It produces the evidence behind the structural claims in
README.md and docs/DATA_QUALITY_COEFFICIENTS.md, so that anyone can re-derive them
instead of taking them on trust.

    uv run python scripts/probe_workbook.py

Every number it prints is reproducible from sources/ on a clean machine.
"""

from __future__ import annotations

import collections
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from import_coeficienti import classify_columns  # noqa: E402

WORKBOOK = Path("sources/Proiect-COEFICIENTI-1-8-MMFTSS-16.07.2026-1000.xlsx")

# Function codes look like 81.10101002.02.2 — family, chapter path, variant, level.
CODE = re.compile(r"^\d{2}\.\d{8}\.\d{2}\.\d$")

# Head nouns that start a Romanian public-sector job title. Used only to decide whether
# a fragment of a merged title cell is plausibly a job title at all; a fragment that
# fails this test is not discarded, it is flagged for human review.
HEAD = (
    r"(director|șef|şef|inspector|expert|asistent|tehnician|consilier|referent|medic|"
    r"comisar|contabil|inginer|auditor|trezorier|secretar|manager|profesor|educator|"
    r"redactor|arhivist|bibliotecar|muzeograf|farmacist|biolog|chimist|fizician|psiholog|"
    r"kinetoterapeut|moaşă|moasa|registrator|statistician|operator|casier|magaziner|portar|"
    r"îngrijitor|ingrijitor|muncitor|şofer|sofer|paznic|analist|programator|administrator|"
    r"arhitect|cercetător|cercetator|procuror|judecător|judecator|grefier|agent|controlor|"
    r"revizor|preot|antrenor|instructor|maistru|laborant|infirmier|brancardier|dietetician|"
    r"logoped|optician|cosmetician|institutor|licenţiat|licentiat)"
)
HAS_WORD = re.compile(r"[A-Za-zĂÂÎȘȚăâîșț]{4}")


def decimals(x: float) -> int:
    """Decimal places in the shortest exact repr. 16 means 'straight out of a division'."""
    s = repr(float(x))
    if "e" in s or "E" in s:
        return 17
    s = s.rstrip("0").rstrip(".")
    return len(s.split(".")[1]) if "." in s else 0


def coded_rows(wb) -> list[tuple[str, int, str, tuple]]:
    """Rows carrying a function code, with their best-guess title cell."""
    out = []
    for name in wb.sheetnames:
        for r, row in enumerate(wb[name].iter_rows(values_only=True), 1):
            if not any(isinstance(v, str) and CODE.match(v.strip()) for v in row):
                continue
            words = [
                v.strip()
                for v in row
                if isinstance(v, str) and v.strip() and not CODE.match(v.strip()) and HAS_WORD.search(v)
            ]
            if words:
                out.append((name, r, max(words, key=len), row))
    return out


def probe_precision(wb) -> None:
    """Distinct coefficient values by decimal place.

    Column selection is delegated to the importer's classifier rather than repeated
    here. An earlier version of this probe carried its own rule — "the column must
    reach 1.5" — which silently discarded three whole sheets whose coefficients never
    get that high, and quoted a distinct-value count the importer disagreed with. One
    definition, one number.
    """
    values: set[float] = set()
    for name in wb.sheetnames:
        rows = [tuple(r) for r in wb[name].iter_rows(values_only=True)]
        roles = classify_columns(rows)
        wanted = {c for c, role in roles.items() if role == "coefficient"}
        for row in rows:
            for c in wanted:
                # int, not just float: the grid floor of 1 is stored as an integer,
                # and testing only for float quietly drops the bottom of the scale.
                value = row[c] if c < len(row) else None
                if isinstance(value, (int, float)) and not isinstance(value, bool) and 0.5 <= value <= 8.5:
                    values.add(float(value))
    hist = collections.Counter(decimals(v) for v in values)
    total = len(values)

    print(f"distinct coefficient values: {total}")
    for k in sorted(hist):
        print(f"  {k:2d} dp -> {hist[k]:4d}  ({hist[k] / total:5.1%})")
    print(f"  <=2 dp: {sum(hist[k] for k in hist if k <= 2)}")
    print(f" >=14 dp: {sum(hist[k] for k in hist if k >= 14)}")
    print(f"  span: {min(values)} .. {max(values)}  ratio 1:{max(values) / min(values):.2f}")
    print("  NB: the maximum is Annex IX's 2030 column. Annex IX phases dignitary")
    print("      coefficients 2026/2027 -> 2031, so the span in force during 2027 is")
    print("      1,00 .. 6,4702 = 1:6,47, and 8,00 is reached only in 2031.")


def probe_assimilation(wb) -> None:
    """How many positions merge former job titles, and by which separator."""
    rows = coded_rows(wb)
    semi = [x for x in rows if ";" in x[2]]
    comma = [x for x in rows if ";" not in x[2] and re.search(r",\s*" + HEAD, x[2].lower())]
    slash = [x for x in rows if ";" not in x[2] and "/" in x[2]]

    clean = review = titles = 0
    for _, _, title, _ in semi:
        frags = [t.strip(" .;,") for t in title.split(";") if len(t.strip(" .;,")) > 2]
        if len(frags) < 2:
            continue
        if all(re.match(HEAD, f.lower()) for f in frags):
            clean += 1
            titles += len(frags)
        else:
            review += 1

    print(f"\nrows carrying a function code: {len(rows)}")
    print(f"  merged with ';' : {len(semi):4d}  ({clean} parse cleanly -> {titles} titles, {review} need review)")
    print(f"  merged with ',' : {len(comma):4d}")
    print(f"  merged with '/' : {len(slash):4d}")
    merged = len(semi) + len(comma) + len(slash)
    print(f"  at least one merge separator: {merged} ({merged / len(rows):.1%} of coded rows)")
    print("\nNo separator is used consistently. ';' also appears inside a single title")
    print("('tehnician superior de imagistică; radiologie; radioterapie'), ',' also introduces")
    print("qualifiers, and '/' also appears inside one title ('secretar instituție/unitate de")
    print("învățământ'). The split is a claim, not a fact — hence assimilation.parse.")


def main() -> None:
    if not WORKBOOK.exists():
        raise SystemExit(f"missing {WORKBOOK}")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    print(f"sheets: {len(wb.sheetnames)}")
    probe_precision(wb)
    probe_assimilation(wb)


if __name__ == "__main__":
    main()
