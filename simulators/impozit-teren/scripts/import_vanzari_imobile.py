"""How many properties changed hands, per county and month — the denominator for a price.

`import_transfer_imobiliar.py` reads a tax and turns it into money: every commune's half of the
art. 111 transfer tax, which across the country implies somewhere between 36 and 107 mld lei of
declared property value in a year. What it cannot say is what a *property* costs, because a sum
of money is not a price until it is divided by a number of things.

ANCPI publishes that number. Every month it files, per county, how many sales were registered,
split by what was sold — and it puts the files on data.gov.ro as open data, which is the
cheapest source in this repository by a wide margin: no crawl, no key, no negotiation.

    preț mediu declarat = impozit ÷ cotă ÷ numărul de vânzări

**`fara constructii` is the line that matters.** A sale with no building on the parcel is a land
sale, and it is the only place in any Romanian public dataset where bare land is counted
separately from land-with-a-house. Two thirds of registered sales are of that kind.

**The six categories are not a partition and must never be summed.** This was measured before
it was written down, because the names invite the opposite assumption. Across 2 659
county-months where all three appear, `agricol + neagricol` equals `fara constructii` exactly
twice, comes within 5% eighty-seven times, and disagrees the other 2 570. Nationally the pair
adds to **41%** of the bare-land count. So the file carries the two axes side by side, as
reported, and the schema says they do not reconcile. Whatever ANCPI is counting under
`agricol`, it is not a decomposition of the land sales.

**And the months do not make a year.** No file exists for 2015 or for any month of 2021; 2019
has seven months, 2020 has eight, 2024 has one; the fullest years have eleven. Summing what
exists and calling it an annual figure would understate 2019 by a third against 2018 for a
reason that is about ANCPI's publishing and not about the property market. Every row therefore
carries `monthsReported`, and `assumptions.monthsByYear` says exactly which months are behind
it, so a reader comparing two years can divide first.

**Forty-seven counties are forty-two.** The county name is spelled several ways across nine
years of files — `BISTRITA NASAUD`, `BISTRITA?NASAUD` where a diacritic died in transit, and
`BISTRITA–NASAUD` with an en-dash; `SATU MARE` and `SATU\xa0MARE` with a non-breaking space —
so the name is folded to letters before it is matched, and a county that fails to fold onto a
known code stops the import rather than quietly becoming a forty-third.

Not wired into CI. It reads a third-party portal and the historical files are occasionally
re-issued, the same reason the other network importers are not diffed.

Usage:
    uv run --with openpyxl python simulators/impozit-teren/scripts/import_vanzari_imobile.py
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
import unicodedata
import urllib.parse
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import politete  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
PORTAL = "https://data.gov.ro"
DATASET = "Dinamica vânzărilor de imobile"

# An open data portal exists to be read in bulk, so the five-second floor that is right for a
# commercial listing site would be theatre here. One second, still one request at a time.
DELAY = 1.0

# ANCPI's labels, and the names this file gives them. `-` is carried because it is a fifth of
# the rows and dropping it would be an editorial decision; it is not a total — Alba filed 184
# under it in February 2016 while filing 213 under `cu constructii` — and ANCPI does not
# document what it means, so it keeps a name that claims nothing.
TYPES = {
    "-": "unspecified",
    "cu constructii": "withBuildings",
    "fara constructii": "withoutBuildings",
    "agricol": "agricultural",
    "neagricol": "nonAgricultural",
    "apartamente": "apartments",
}

COUNTY_CODE = {
    "alba": "AB", "arad": "AR", "arges": "AG", "bacau": "BC", "bihor": "BH",
    "bistritanasaud": "BN", "botosani": "BT", "braila": "BR", "brasov": "BV",
    "bucuresti": "B", "buzau": "BZ", "calarasi": "CL", "carasseverin": "CS", "cluj": "CJ",
    "constanta": "CT", "covasna": "CV", "dambovita": "DB", "dolj": "DJ", "galati": "GL",
    "giurgiu": "GR", "gorj": "GJ", "harghita": "HR", "hunedoara": "HD", "ialomita": "IL",
    "iasi": "IS", "ilfov": "IF", "maramures": "MM", "mehedinti": "MH", "mures": "MS",
    "neamt": "NT", "olt": "OT", "prahova": "PH", "salaj": "SJ", "satumare": "SM",
    "sibiu": "SB", "suceava": "SV", "teleorman": "TR", "timis": "TM", "tulcea": "TL",
    "valcea": "VL", "vaslui": "VS", "vrancea": "VN",
}

COUNTY_NAMES = {
    "AB": "Alba", "AG": "Argeș", "AR": "Arad", "B": "București", "BC": "Bacău", "BH": "Bihor",
    "BN": "Bistrița-Năsăud", "BR": "Brăila", "BT": "Botoșani", "BV": "Brașov", "BZ": "Buzău",
    "CJ": "Cluj", "CL": "Călărași", "CS": "Caraș-Severin", "CT": "Constanța", "CV": "Covasna",
    "DB": "Dâmbovița", "DJ": "Dolj", "GJ": "Gorj", "GL": "Galați", "GR": "Giurgiu",
    "HD": "Hunedoara", "HR": "Harghita", "IF": "Ilfov", "IL": "Ialomița", "IS": "Iași",
    "MH": "Mehedinți", "MM": "Maramureș", "MS": "Mureș", "NT": "Neamț", "OT": "Olt",
    "PH": "Prahova", "SB": "Sibiu", "SJ": "Sălaj", "SM": "Satu Mare", "SV": "Suceava",
    "TL": "Tulcea", "TM": "Timiș", "TR": "Teleorman", "VL": "Vâlcea", "VN": "Vrancea",
    "VS": "Vaslui",
}


def fold(name: str) -> str:
    """Letters only. Nine years of files spell the same county five different ways."""
    decomposed = unicodedata.normalize("NFD", name.replace("\xa0", " ").lower())
    stripped = "".join(c for c in decomposed if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z]", "", stripped)


def decode(raw: bytes) -> str:
    """The CSVs are windows-1250 and the newer ones are UTF-8, with no way to tell but trying."""
    for encoding in ("utf-8-sig", "utf-8", "windows-1250", "iso-8859-2", "cp1252"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", "replace")


def table(raw: bytes, is_spreadsheet: bool) -> list[list[str]]:
    """Rows, from either an XLSX or a CSV whose delimiter is a comma or a semicolon."""
    if is_spreadsheet:
        import openpyxl

        book = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = book[book.sheetnames[0]]
        return [
            ["" if cell is None else str(cell) for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    text = decode(raw)
    first = text.splitlines()[0] if text.splitlines() else ""
    delimiter = ";" if first.count(";") > first.count(",") else ","
    return list(csv.reader(io.StringIO(text), delimiter=delimiter))


def period(value: str) -> tuple[int, int] | None:
    """`01.02.2016` and `2016-02-01` both appear. Returns (year, month) of the report."""
    dotted = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})", value)
    if dotted:
        return int(dotted.group(3)), int(dotted.group(2))
    iso = re.search(r"(\d{4})-(\d{2})-(\d{2})", value)
    if iso:
        return int(iso.group(1)), int(iso.group(2))
    return None


def resources(fetcher: politete.Politete) -> list[str]:
    """Every file of this dataset family, discovered through pages the portal permits.

    The obvious way to do this is CKAN's `/api/3/action/package_search`, and it is the wrong
    way: data.gov.ro's robots.txt says `Disallow: /api/`. That rule is stock CKAN and is
    plainly aimed at search engines rather than at people fetching open data — but a gate that
    is obeyed only when it seems well-aimed is not a gate, and `politete.get` raised
    `Disallowed` rather than let this file decide it knew better.

    So discovery goes through `/dataset?q=...` and each dataset's own page, which the same file
    permits, and which carry the identical `/download/` links. It costs fifteen extra requests
    at the portal's stated ten-second `Crawl-Delay`, and it is the difference between reading
    a public catalogue the way it asks to be read and the way that happened to be convenient.
    """
    listing = fetcher.get(f"{PORTAL}/dataset?q=%22{urllib.parse.quote(DATASET)}%22")
    pages = sorted(set(re.findall(r'href="(/dataset/[a-z0-9-]+)"', listing)))
    if not pages:
        raise SystemExit("no dataset pages matched; the portal's markup may have changed")
    found: list[str] = []
    for page in pages:
        html = fetcher.get(f"{PORTAL}{page}")
        found.extend(
            re.findall(r'href="(https://data\.gov\.ro/dataset/[^"]*?/download/[^"]+)"', html)
        )
    return sorted(set(found))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--delay", type=float, default=DELAY)
    args = parser.parse_args()

    fetcher = politete.Politete(
        cache=ROOT / "sources" / "ancpi", delay=args.delay, budget=400
    )
    files = resources(fetcher)
    print(f"{len(files)} files on data.gov.ro", file=sys.stderr)

    # (county, year, month) -> {name: count}
    cells: dict[tuple[str, int, int], dict[str, int]] = defaultdict(dict)
    months: dict[int, set[int]] = defaultdict(set)
    unknown_counties: set[str] = set()
    unknown_types: set[str] = set()

    for index, url in enumerate(files, 1):
        raw = fetcher.get_bytes(url)
        rows = table(raw, url.lower().endswith((".xlsx", ".xls")))
        if not rows:
            continue
        header = [str(cell).strip().upper() for cell in rows[0]]
        at = {name: position for position, name in enumerate(header)}
        needed = ("JUDET", "LUNA_RAPORTATA", "TIP_PROPRIETATE", "VANZARI")
        if not all(name in at for name in needed):
            continue
        for row in rows[1:]:
            if len(row) <= max(at[name] for name in needed):
                continue
            folded = fold(str(row[at["JUDET"]]))
            if not folded:
                continue
            code = COUNTY_CODE.get(folded)
            if not code:
                unknown_counties.add(folded)
                continue
            when = period(str(row[at["LUNA_RAPORTATA"]]))
            if not when:
                continue
            label = str(row[at["TIP_PROPRIETATE"]]).strip().lower()
            name = TYPES.get(label)
            if not name:
                unknown_types.add(label)
                continue
            try:
                count = int(float(str(row[at["VANZARI"]]).strip().replace(" ", "")))
            except ValueError:
                continue
            cells[(code, when[0], when[1])][name] = count
            months[when[0]].add(when[1])
        if index % 20 == 0:
            print(f"  {index}/{len(files)}", file=sys.stderr)

    # A county that will not fold is a county silently missing from every total below it.
    if unknown_counties:
        raise SystemExit(f"county names that did not fold onto a code: {sorted(unknown_counties)}")
    if unknown_types:
        raise SystemExit(f"property types nobody has named: {sorted(unknown_types)}")

    by_county: dict[str, dict[int, dict[str, int]]] = defaultdict(lambda: defaultdict(dict))
    reported: dict[tuple[str, int], set[int]] = defaultdict(set)
    for (code, year, month), counts in cells.items():
        reported[(code, year)].add(month)
        for name, count in counts.items():
            by_county[code][year][name] = by_county[code][year].get(name, 0) + count

    counties = []
    for code in sorted(by_county, key=lambda c: COUNTY_NAMES.get(c, c)):
        series = []
        for year in sorted(by_county[code]):
            series.append(
                {
                    "year": year,
                    "monthsReported": len(reported[(code, year)]),
                    "sales": {
                        name: by_county[code][year][name]
                        for name in sorted(by_county[code][year])
                    },
                }
            )
        counties.append({"county": code, "name": COUNTY_NAMES[code], "series": series})

    # The measured disagreement, computed rather than asserted, because the limitation that
    # says these do not reconcile is only worth having if the number behind it is in the file.
    pair = land = 0
    for counts in cells.values():
        if {"agricultural", "nonAgricultural", "withoutBuildings"} <= set(counts):
            pair += counts["agricultural"] + counts["nonAgricultural"]
            land += counts["withoutBuildings"]

    # Ties go to the later year. Five years have eleven months each, and picking whichever the
    # dictionary happened to yield first put 2017 in the summary of a file that runs to 2024 —
    # true, useless, and the kind of thing a reader takes as a statement about recency.
    fullest = max(months, key=lambda y: (len(months[y]), y))
    national = defaultdict(int)
    for county in counties:
        for row in county["series"]:
            if row["year"] == fullest:
                for name, count in row["sales"].items():
                    national[name] += count

    document = {
        "$schema": "../schema/vanzari-imobile.schema.json",
        "id": f"vanzari-imobile-{max(months)}",
        "title": (
            f"Numărul vânzărilor de imobile înregistrate, pe județe, "
            f"{min(months)}–{max(months)}"
        ),
        "publisher": "ANCPI",
        "period": f"{min(months)}-{max(months)}",
        "currency": "RON",
        "provenance": {
            "source": "ancpi-dinamica-vanzarilor-de-imobile",
            "locator": (
                f'data.gov.ro, seturile „{DATASET}" ale ANCPI, {len(files)} fișiere lunare '
                "CSV și XLSX, coloanele JUDET, LUNA_RAPORTATA, TIP_PROPRIETATE, VANZARI"
            ),
            "confidence": "verbatim",
            "note": (
                "Numerele sunt cele raportate de ANCPI, însumate pe an în acest import. Nu se "
                "calculează nimic altceva: nu există valori și nici suprafețe în sursă, doar "
                "câte vânzări s-au înregistrat."
            ),
        },
        "assumptions": {
            "monthsByYear": {str(year): sorted(found) for year, found in sorted(months.items())},
            "types": TYPES,
            "fullestYear": fullest,
            "note": (
                "Anii nu sunt completi și doi lipsesc de tot. Însumarea lunilor existente nu "
                "dă un an; „monthsReported” spune peste câte luni s-a adunat fiecare rând, "
                "iar o comparație între ani trebuie să împartă mai întâi."
            ),
        },
        "summary": {
            "counties": len(counties),
            "files": len(files),
            "years": len(months),
            "fullestYear": fullest,
            "salesInFullestYear": dict(sorted(national.items())),
            "agriculturalPlusNonAgriculturalOverWithoutBuildings": (
                round(pair / land, 4) if land else None
            ),
            "crawl": fetcher.report(),
        },
        "counties": counties,
        "limitations": [
            {
                "id": "categoriile-nu-se-aduna",
                "severity": "blocking",
                "text": (
                    "Cele șase categorii nu sunt o împărțire a aceluiași total și nu se pot "
                    "aduna. Măsurat pe toate lunile în care apar toate trei, „agricol” plus "
                    "„neagricol” face "
                    f"{round(100 * pair / land, 1) if land else '—'}% din „fără construcții”, "
                    "nu 100%. Sunt două axe de clasificare diferite, publicate una lângă alta; "
                    "ANCPI nu documentează relația dintre ele."
                ),
                "affects": ["vanzari-imobile", "transfer-imobiliar"],
            },
            {
                "id": "lunile-nu-fac-un-an",
                "severity": "blocking",
                "text": (
                    "Nu există niciun fișier pentru 2015 și niciunul pentru vreo lună din 2021; "
                    "2019 are șapte luni, 2020 opt, 2024 una, iar anii cei mai plini au "
                    "unsprezece. Suma lunilor existente nu este un an. Fiecare rând poartă "
                    "„monthsReported”, iar „monthsByYear” spune exact ce luni stau în spate."
                ),
                "affects": ["vanzari-imobile"],
            },
            {
                "id": "clasificarea-s-a-schimbat-in-2018",
                "severity": "blocking",
                "text": (
                    "În 2018 ANCPI a schimbat modul de clasificare, iar seria nu este continuă "
                    "peste acel an. „Apartamente” cade de la 111 028 de vânzări în 2017 la 20 "
                    "în 2018, „-” de la 103 334 la 1 543, în timp ce „fără construcții” aproape "
                    "se dublează, de la 212 831 la 392 512. Aceleași dosare, altă etichetă. "
                    "Suma „cu construcții” plus „fără construcții” acoperă deci aproape tot "
                    "după 2018 și îi lipsesc vreo 230 000 de vânzări pe an înainte — un preț "
                    "mediu calculat pe anii 2016 sau 2017 iese cu vreo 40% prea mare. Anii de "
                    "dinainte de 2018 nu se compară cu cei de după."
                ),
                "affects": ["vanzari-imobile", "pret-tranzactie"],
            },
            {
                "id": "numere-nu-valori",
                "severity": "material",
                "text": (
                    "Sursa numără vânzări, nu bani și nu hectare. Un preț mediu iese doar "
                    "împărțind valoarea declarată din impozitul pe transfer la numărul de aici, "
                    "iar cele două nu acoperă exact aceleași tranzacții: impozitul nu se "
                    "datorează la moșteniri, donații între rude și reconstituiri de "
                    "proprietate, care totuși se înregistrează în cartea funciară."
                ),
                "affects": ["vanzari-imobile", "transfer-imobiliar"],
            },
            {
                "id": "categoria-fara-nume",
                "severity": "material",
                "text": (
                    "O cincime din rânduri sunt depuse sub tipul „-”, pe care ANCPI nu îl "
                    "explică. Nu este un total: în februarie 2016 Alba a depus 184 sub „-” și "
                    "213 sub „cu construcții”. Este păstrat ca „unspecified”, cu numele care "
                    "nu pretinde nimic."
                ),
                "affects": ["vanzari-imobile"],
            },
            {
                "id": "judetul-scris-in-cinci-feluri",
                "severity": "note",
                "text": (
                    "Numele județului apare în cinci ortografii peste nouă ani — cu cratimă, cu "
                    "linie de dialog, cu diacritice pierdute și cu spațiu neîntrerupt. Sunt "
                    "reduse la litere înainte de potrivire, iar un nume care nu se potrivește "
                    "oprește importul în loc să devină al patruzeci și treilea județ."
                ),
                "affects": ["vanzari-imobile"],
            },
        ],
    }

    out = ROOT / "data" / f"vanzari-imobile-{max(months)}.json"
    out.write_text(json.dumps(document, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"\n{fetcher.report()}")
    print(f"{len(counties)} județe, {len(months)} ani, cel mai plin {fullest}")
    print(f"  vânzări în {fullest}: {dict(sorted(national.items()))}")
    if land:
        print(f"  agricol+neagricol / fără construcții = {pair / land:.3f}")
    print(f"Wrote {out.relative_to(ROOT.parents[1])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
